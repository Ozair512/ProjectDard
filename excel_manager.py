"""
Excel Manager for Project Dard
Handles automatic Excel file creation and data export
"""

import os
import pandas as pd
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

class ExcelManager:
    def __init__(self, excel_dir='excel_exports'):
        self.excel_dir = excel_dir
        self.ensure_directory_exists()
        
    def ensure_directory_exists(self):
        """Create excel exports directory if it doesn't exist"""
        if not os.path.exists(self.excel_dir):
            os.makedirs(self.excel_dir)
    
    def get_excel_path(self, filename):
        """Get full path for Excel file"""
        return os.path.join(self.excel_dir, filename)
    
    def create_or_update_topics_excel(self, topics_data):
        """Create or update topics Excel file"""
        filename = 'weekly_topics.xlsx'
        filepath = self.get_excel_path(filename)
        
        # Convert to DataFrame
        df = pd.DataFrame(topics_data)
        
        if os.path.exists(filepath):
            # Load existing workbook
            with pd.ExcelWriter(filepath, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                df.to_excel(writer, sheet_name='Topics', index=False)
        else:
            # Create new workbook
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Topics', index=False)
        
        self.format_excel_file(filepath, 'Topics')
        return filepath
    
    def create_or_update_stories_excel(self, stories_data):
        """Create or update stories Excel file"""
        filename = 'weekly_stories.xlsx'
        filepath = self.get_excel_path(filename)
        
        # Convert to DataFrame
        df = pd.DataFrame(stories_data)
        
        if os.path.exists(filepath):
            # Load existing workbook
            with pd.ExcelWriter(filepath, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                df.to_excel(writer, sheet_name='Stories', index=False)
        else:
            # Create new workbook
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Stories', index=False)
        
        self.format_excel_file(filepath, 'Stories')
        return filepath
    
    def create_or_update_comments_excel(self, comments_data):
        """Create or update comments Excel file"""
        filename = 'comments.xlsx'
        filepath = self.get_excel_path(filename)
        
        # Convert to DataFrame
        df = pd.DataFrame(comments_data)
        
        if os.path.exists(filepath):
            # Load existing workbook
            with pd.ExcelWriter(filepath, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                df.to_excel(writer, sheet_name='Comments', index=False)
        else:
            # Create new workbook
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Comments', index=False)
        
        self.format_excel_file(filepath, 'Comments')
        return filepath
    
    def create_or_update_appeals_excel(self, appeals_data):
        """Create or update appeals Excel file"""
        filename = 'appeals.xlsx'
        filepath = self.get_excel_path(filename)
        
        # Convert to DataFrame
        df = pd.DataFrame(appeals_data)
        
        if os.path.exists(filepath):
            # Load existing workbook
            with pd.ExcelWriter(filepath, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                df.to_excel(writer, sheet_name='Appeals', index=False)
        else:
            # Create new workbook
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Appeals', index=False)
        
        self.format_excel_file(filepath, 'Appeals')
        return filepath
    
    def create_or_update_visitors_excel(self, visitors_data):
        """Create or update visitors Excel file"""
        filename = 'visitors.xlsx'
        filepath = self.get_excel_path(filename)
        
        # Convert to DataFrame
        df = pd.DataFrame(visitors_data)
        
        if os.path.exists(filepath):
            # Load existing workbook
            with pd.ExcelWriter(filepath, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                df.to_excel(writer, sheet_name='Visitors', index=False)
        else:
            # Create new workbook
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Visitors', index=False)
        
        self.format_excel_file(filepath, 'Visitors')
        return filepath
    
    def format_excel_file(self, filepath, sheet_name):
        """Format Excel file with proper styling"""
        try:
            wb = load_workbook(filepath)
            ws = wb[sheet_name]
            
            # Header formatting
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # Apply header formatting
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(filepath)
        except Exception as e:
            print(f"Error formatting Excel file {filepath}: {str(e)}")
    
    def export_all_data(self):
        """Export all database data to Excel files"""
        from app import app, db
        
        with app.app_context():
            from models import WeeklyTopic, WeeklyStory, Comment, Appeal, Visitor
            
            # Topics
            topics = WeeklyTopic.query.all()
            topics_data = []
            for topic in topics:
                topics_data.append({
                    'ID': topic.id,
                    'Title (English)': topic.title_en,
                    'Title (Persian)': topic.title_fa,
                    'Content (English)': topic.content_en[:500] + '...' if len(topic.content_en) > 500 else topic.content_en,
                    'Content (Persian)': topic.content_fa[:500] + '...' if topic.content_fa and len(topic.content_fa) > 500 else topic.content_fa,
                    'Created At': topic.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                    'Is Active': topic.is_active,
                    'Admin ID': topic.admin_id
                })
            
            # Stories
            stories = WeeklyStory.query.all()
            stories_data = []
            for story in stories:
                stories_data.append({
                    'ID': story.id,
                    'Title (English)': story.title_en,
                    'Title (Persian)': story.title_fa,
                    'Content (English)': story.content_en[:500] + '...' if len(story.content_en) > 500 else story.content_en,
                    'Content (Persian)': story.content_fa[:500] + '...' if story.content_fa and len(story.content_fa) > 500 else story.content_fa,
                    'Author (English)': story.author_en,
                    'Author (Persian)': story.author_fa,
                    'Created At': story.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                    'Is Active': story.is_active,
                    'Admin ID': story.admin_id
                })
            
            # Comments
            comments = Comment.query.all()
            comments_data = []
            for comment in comments:
                comments_data.append({
                    'ID': comment.id,
                    'Content': comment.content[:300] + '...' if len(comment.content) > 300 else comment.content,
                    'Language': comment.language,
                    'Author Name': comment.author_name,
                    'Author Email': comment.author_email,
                    'Created At': comment.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                    'Is Hidden': comment.is_hidden,
                    'Likes': comment.likes,
                    'Dislikes': comment.dislikes,
                    'Topic ID': comment.topic_id
                })
            
            # Appeals
            appeals = Appeal.query.all()
            appeals_data = []
            for appeal in appeals:
                appeals_data.append({
                    'ID': appeal.id,
                    'Comment ID': appeal.comment_id,
                    'Appellant Name': appeal.appellant_name,
                    'Appellant Email': appeal.appellant_email,
                    'Reason': appeal.reason[:300] + '...' if len(appeal.reason) > 300 else appeal.reason,
                    'Status': appeal.status,
                    'Created At': appeal.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                    'Reviewed At': appeal.reviewed_at.strftime('%Y-%m-%d %H:%M:%S') if appeal.reviewed_at else '',
                    'Admin ID': appeal.admin_id
                })
            
            # Visitors
            visitors = Visitor.query.all()
            visitors_data = []
            for visitor in visitors:
                visitors_data.append({
                    'ID': visitor.id,
                    'IP Address': visitor.ip_address,
                    'User Agent': visitor.user_agent[:100] + '...' if visitor.user_agent and len(visitor.user_agent) > 100 else visitor.user_agent,
                    'Page Visited': visitor.page_visited,
                    'Visited At': visitor.visited_at.strftime('%Y-%m-%d %H:%M:%S')
                })
            
            # Create Excel files
            exported_files = []
            if topics_data:
                exported_files.append(self.create_or_update_topics_excel(topics_data))
            if stories_data:
                exported_files.append(self.create_or_update_stories_excel(stories_data))
            if comments_data:
                exported_files.append(self.create_or_update_comments_excel(comments_data))
            if appeals_data:
                exported_files.append(self.create_or_update_appeals_excel(appeals_data))
            if visitors_data:
                exported_files.append(self.create_or_update_visitors_excel(visitors_data))
            
            return exported_files

# Global instance
excel_manager = ExcelManager()